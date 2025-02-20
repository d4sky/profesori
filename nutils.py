import sys, os

import numpy as np
import pandas as pd
import random
import string

from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

colorsDict = {
    "black": "000000",
    "white": "FFFFFF",
    "red": "FF0000",
    "green": "00FF00",
    "blue": "0000FF",
    "yellow": "FFFF00",
    "magenta": "FF00FF",
    "cyan": "00FFFF",
    "brown": "A52A2A",  # Brown
    "orange": "FFA500",
    # Add more colors as needed
}

def modify_filename(filename, insLength=3):
    """
    Modify the filename according to the given rules.
    """
    if len(filename) < 5:
        S1 = ''
        S2 = filename
    else:
        S1 = filename[:-5]
        S2 = filename[-5:]

    inS = ''.join(random.choice(string.ascii_letters) for _ in range(insLength))
    if S1: inS = '_' + inS

    newfilename = S1 + inS + S2
    return newfilename
    
class FormatMatrix:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.general_format_info = {
            "size": 10, 
            "bold": False, 
            "fontColor": colorsDict["black"], 
            "fgColor": colorsDict["white"], 
            "horizontal": "center", 
            "vertical": "center",
            "border": "thin", 
            "brColor": colorsDict["black"]
        }

    def Set_default_formatting(self, size=10, bold=False, fontColor="black", fgColor="white", horizontal="center", vertical="center", border="thin", brColor="black"):
        self.general_format_info = {
            "size": size, 
            "bold": bold, 
            "fontColor": colorsDict[fontColor], 
            "fgColor": colorsDict[fgColor], 
            "horizontal": horizontal, 
            "vertical": vertical,
            "border": border, 
            "brColor": colorsDict[brColor]
        }

    def Apply_default_formatting(self):
        """
        Apply the default formatting to all cells in the active worksheet.
        """
        font = Font(size=self.general_format_info["size"], 
                    bold=self.general_format_info["bold"], 
                    color=self.general_format_info["fontColor"])

        fill = PatternFill(start_color=self.general_format_info["fgColor"], 
                           end_color=self.general_format_info["fgColor"], 
                           fill_type="solid")

        alignment = Alignment(horizontal=self.general_format_info["horizontal"], 
                              vertical=self.general_format_info["vertical"])

        border = Border(
            left=Side(style=self.general_format_info["border"], color=self.general_format_info["brColor"]),
            right=Side(style=self.general_format_info["border"], color=self.general_format_info["brColor"]),
            top=Side(style=self.general_format_info["border"], color=self.general_format_info["brColor"]),
            bottom=Side(style=self.general_format_info["border"], color=self.general_format_info["brColor"])
        )

        for row in self.ws.iter_rows():
            for cell in row:
                cell.font = font
                cell.fill = fill
                cell.alignment = alignment
                cell.border = border

    def Apply_border_formatting(self, row_beg, row_end, col_beg, col_end, border_thickness = 'thick', border_color='black'):
        if border_color == "black":
          new_side = Side(style=border_thickness)
        else:
          new_side = Side(style=border_thickness, color=colorsDict[border_color])
        
        for row in range(row_beg, row_end + 1):
            for col in range(col_beg, col_end + 1):
                cell = self.ws.cell(row=row, column=col)
                current_border = cell.border
                if col_beg == 1 and col_end == 1:
                  new_border = Border(
                    top=new_side,
                    bottom=new_side,
                    left=new_side if col == col_beg else current_border.left,
                    right=new_side if col == col_end else current_border.right
                  )
                elif row_beg == 1 and row_end == 1:  
                  new_border = Border(
                    top=new_side if row == row_beg else current_border.top,
                    bottom=new_side if row == row_end else current_border.bottom,
                    left=new_side,
                    right=new_side
                  )
                else:
                  new_border = Border(
                    top=new_side if row == row_beg else current_border.top,
                    bottom=new_side if row == row_end else current_border.bottom,
                    left=new_side if col == col_beg else current_border.left,
                    right=new_side if col == col_end else current_border.right
                  )
                cell.border = new_border                
    
    '''
    def Merge_cells(self, start_row, end_row, start_col, end_col, value=None):
        self.ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
        top_left_cell = self.ws.cell(row=start_row, column=start_col)
        top_left_cell.value = top_left_cell.value[:-2]
        if value is not None:
            top_left_cell.value = value
    '''

    def Merge_cells(self, start_row, end_row, start_col, end_col, value=None, font_color="black", bg_color="white"):
        self.ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
        top_left_cell = self.ws.cell(row=start_row, column=start_col)
        top_left_cell.value = value if value is not None else top_left_cell.value

        # Change font color and background color
        top_left_cell.font = Font(color=colorsDict[font_color])
        top_left_cell.fill = PatternFill(start_color=colorsDict[bg_color], end_color=colorsDict[bg_color], fill_type="solid")

        # Apply formatting to all merged cells
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = self.ws.cell(row=row, column=col)
                cell.font = Font(color=colorsDict[font_color])
                cell.fill = PatternFill(start_color=colorsDict[bg_color], end_color=colorsDict[bg_color], fill_type="solid")


    def Apply_group_bordering(self, chains, properties):
        def Get_color(rowi, coli, chaini):
          color = "black"
          label = self.ws.cell(row=rowi, column=coli).value
          parts = label.split('_') if label is not None else []
          if len(parts) == 2: 
            resi = parts[0]
            propID = f"{chains[0][chaini]}_{resi}"
            if propID in properties:
              color = "red"
              
          return color
        
        Nrows = self.ws.max_row
        Ncols = self.ws.max_column
        
        # Apply bordering on each cell in Leader
        i = 2
        while i<=Nrows:
            new_color = Get_color(i, 1, 0)
            self.Apply_border_formatting( i, i+1, 1, 1)
            #self.Apply_border_formatting( i, i+1, 1, 1, border_color=new_color)
            #self.Merge_cells( i, i+1, 1, 1)
            self.Merge_cells( i, i+1, 1, 1, font_color = new_color)
            i += 2

        # Apply bordering on each cell in Header
        j = 2
        while j<=Ncols:
            new_color = Get_color(1, j, 1)
            #self.Apply_border_formatting( 1, 1, j, j+1, border_color=new_color)
            self.Apply_border_formatting( 1, 1, j, j+1)
            #self.Merge_cells( 1, 1, j, j+1)
            self.Merge_cells( 1, 1, j, j+1, font_color = new_color)
            j += 2

        # Apply bordering on 2x2 Sub-Cells
        i = 2
        while i<=Nrows:
            j = 2
            while j<Ncols:
                self.Apply_border_formatting(i, i+1, j, j+1)
                j += 2
            i += 2
            
    def Apply_prop_formatting(self, properties):
        def Apply_format(wcell, backgroundC, fontC = "black"):
            wcell.fill = PatternFill("solid", fgColor=colorsDict[backgroundC])
            wcell.font = Font(color=colorsDict[fontC])
            return True
          
    
        formatted_indexes = []
        for key, prop_data in properties.items():
            # Extract indices from the key
            i, j = map(int, key.split('_'))
            # Adjust indices to account for the additional row and column
            i += 2
            j += 2
            # Apply formatting to the cell
            cell = self.ws.cell(row=i, column=j)
            was_formatted = False
            for prop in prop_data:
              if   prop == 'bb':
                # Assuming prop_data contains formatting information
                '''
                cell.font = Font(...)  # Apply font formatting
                cell.alignment = Alignment(...)  # Apply alignment formatting      
                '''
                #cell.fill = PatternFill(start_color=colorsDict["yellow"], end_color=colorsDict["red"], fill_type="solid")  # Apply fill formatting
                was_formatted = Apply_format(cell, "blue", "white")
              elif prop == "elec":
                cell.fill = PatternFill("solid", fgColor=colorsDict["red"])
                was_formatted = Apply_format(cell, "red")
              else: 
                if isinstance(prop, (int, float)):
                  if prop <= 8.0:
                    was_formatted = Apply_format(cell, "green")
                  if prop <= 6.0:
                    was_formatted = Apply_format(cell, "orange")
                  if prop <= 4.0:
                    was_formatted = Apply_format(cell, "yellow")
                  if prop <= 3.5:
                    was_formatted = Apply_format(cell, "cyan")
            if was_formatted: 
                formatted_indexes.append((i,j))
        
        Nrows = self.ws.max_row
        Ncols = self.ws.max_column
        i = 1
        for i in range(2, Nrows+1): 
          for j in range(2, Ncols+1):
            if (i,j) in formatted_indexes:
              pass
            else:
              #self.ws.cell(row=i, column=j, value = 1000.0)
              #self.ws.cell(row=i, column=j, value = None)
              self.ws.cell(row=i, column=j, value = "")
              
    def Save(self, filename="out.xlsx"):
        """
        Save the workbook to the specified filename.
        """
        try:
            self.wb.save(filename)
            print(f"Workbook saved successfully as {filename}.")
        except PermissionError:
            print(f"Error: The file '{filename}' is currently open.")
            filename = modify_filename(filename)
            self.wb.save(filename)
            print(f"Workbook saved successfully as new {filename}.")

class GeneralMatrix:
    def __init__(self, ids1, ids2):
        """
        Initialize with two lists of IDs.
        """
        self.ids1 = ids1
        self.ids2 = ids2
        self.matrix = self._initialize_matrix()
        self.props  = self._initialize_matrix("list")

        self.formatting = {}
        self.general_format_info = {
            "size": 10, 
            "bold": False, 
            "fontColor": colorsDict["black"], 
            "fgColor": colorsDict["white"], 
            "horizontal": "center", 
            "vertical": "center"
        }

    def _initialize_matrix(self, wtype = "general"):
        """
        Initialize an empty matrix with dimensions corresponding to the lengths of the ID lists.
        """
        if    wtype == "number":
          return np.empty((len(self.ids1), len(self.ids2)), dtype=float)
        elif  wtype == "list":
          return np.empty((len(self.ids1), len(self.ids2)), dtype=list)
        else:
          return np.empty((len(self.ids1), len(self.ids2)), dtype=object)
        
    def Set_element(self, id1, id2, value, what = "matrix"):
        """
        Set the value in the matrix for the given pair of IDs.
        """
        idx1 = self.ids1.index(id1)
        idx2 = self.ids2.index(id2)
        if   what == "matrix":
            self.matrix[idx1, idx2] = value
        elif what == "props":
            if self.props[idx1, idx2] is None: self.props[idx1, idx2] = []  # Initialize as an empty list if None
            self.props[idx1, idx2].append(value)
        else:
            print("Invalid 'what' parameter. Use 'matrix' or 'props'.")
        
    def Get_Format_Matrix(self):
        """
        Create an instance of FormatMatrix and populate its workbook with data from this matrix.
        """
        format_matrix = FormatMatrix()
        ws = format_matrix.wb.active
        out_props = {}

        # Write the ids1 as header in the first row (starting from the second column)
        for j, id2 in enumerate(self.ids2, start=2):
            ws.cell(row=1, column=j, value=id2.replace("_b",""))

        # Write the ids2 as header in the first column (starting from the second row)
        for i, id1 in enumerate(self.ids1, start=2):
            ws.cell(row=i, column=1, value=id1.replace("_b",""))

        # Write the matrix data starting from (2,2)
        for i in range(len(self.ids1)):
            for j in range(len(self.ids2)):
                ws.cell(row=i+2, column=j+2, value=self.matrix[i, j])
                prop_data = self.props[i, j]
                # Check if props is not empty and store the indices and props data
                if prop_data and prop_data != []:
                    out_props[f"{i}_{j}"] = prop_data

        format_matrix.Apply_default_formatting()
        return format_matrix, out_props

    def set_decimal_places(self, decimal_places):
        """
        Set the number of decimal places for all numerical values in the matrix.

        Parameters:
        decimal_places (int): The number of decimal places to format the numbers to.
        """
        format_str = f"{{:.{decimal_places}f}}"
        for i in range(len(self.ids1)):
            for j in range(len(self.ids2)):
                value = self.matrix[i, j]
                if isinstance(value, (int, float)):
                    self.matrix[i, j] = float(format_str.format(value))
    
    def erase(self, val=0.0):
        """
        Replace all elements with a value of 0 with an empty string ('').
        """
        for i in range(len(self.ids1)):
            for j in range(len(self.ids2)):
                if self.matrix[i, j] >= val:
                    self.matrix[i, j] = ''
    
    def Print(self, column_width=10):
        """
        Print the matrix in a readable format.

        Parameters:
        column_width (int): Width of the columns in the printed matrix.
        """
        header_format = f"{{:>{column_width}}}"
        value_format = f"{{:>{column_width}.2f}}"
        empty_format = f"{{:>{column_width}}}"

        # Print column headers
        print(header_format.format(""), end="")
        for id2 in self.ids2:
            print(header_format.format(id2), end="")
        print()

        # Print matrix values
        for i, id1 in enumerate(self.ids1):
            print(header_format.format(id1), end="")
            for j in range(len(self.ids2)):
                value = self.matrix[i, j]
                if isinstance(value, (int, float)):
                    print(value_format.format(value), end="")
                else:
                    print(empty_format.format(""), end="")
            print()
    
    def save_as_excel(self, file_path):
        """
        Save the GeneralMatrix as an Excel file using the ExcelTable class.

        Parameters:
        file_path (str): The file path where the Excel file will be saved.
        """
        excel_table = ExcelTable()
        excel_table.Save(file_path)

        # Add column headers 
        headers = ['AA'] + self.ids2
        excel_table.Add_empty_columns(headers)
        
        # Create a list to store rows for the Excel table
        excel_rows = []
        # Add data rows
        for i, id1 in enumerate(self.ids1):
            row = [id1]
            for j in range(len(self.ids2)):
                value = self.matrix[i, j]
                if isinstance(value, (int, float)):
                    row.append(value)
                else:
                    row.append('')
            excel_rows.append(row)
        
        # Append rows to the Excel table

        excel_table.Append_rows(excel_rows)
        
    def get_value(self, id1, id2):
        """
        Get the value from the matrix for the given pair of IDs.
        """
        idx1 = self.ids1.index(id1)
        idx2 = self.ids2.index(id2)
        return self.matrix[idx1, idx2]
    
    def get_matrix(self):
        """
        Get the entire matrix.
        """
        return self.matrix
    
    def get_submatrix(self, selected_rows, selected_cols):
        """
        Select a submatrix based on the provided lists of strings for rows and columns.

        Parameters:
        selected_rows (list of str): List of strings to select rows.
        selected_cols (list of str): List of strings to select columns.

        Returns:
        np.array: Submatrix selected from the original matrix.
        """
        row_indices = [self.ids1.index(row) for row in selected_rows]
        col_indices = [self.ids2.index(col) for col in selected_cols]
        
        submatrix = self.matrix[row_indices][:, col_indices]
        
        return submatrix
        
    def create_submatrix(self, selected_rows=[], selected_cols=[], inv1=False, inv2=False):
        """
        Get a submatrix based on the provided lists of strings for rows and columns.
        Return a new instance of the class GeneralMatrix with the submatrix.

        Parameters:
        selected_rows (list of str): List of strings to select rows.
        selected_cols (list of str): List of strings to select columns.
        inv1 (bool): If True, select rows not in selected_rows.
        inv2 (bool): If True, select columns not in selected_cols.
    
        Returns:
        GeneralMatrix: New instance of the class with the submatrix.
        """
        if not selected_rows:
            selected_rows = self.ids1
        if not selected_cols:
            selected_cols = self.ids2

        if not inv1:
            if not all(item in self.ids1 for item in selected_rows):
                return False
        if not inv2:
            if not all(item in self.ids2 for item in selected_cols):
                return False
        
        if inv1:
            row_indices = [i for i, row in enumerate(self.ids1) if row not in selected_rows]
        else:
            row_indices = [self.ids1.index(row) for row in selected_rows]
        
        if inv2:
            col_indices = [i for i, col in enumerate(self.ids2) if col not in selected_cols]
        else:
            col_indices = [self.ids2.index(col) for col in selected_cols]
        
        submatrix = self.matrix[row_indices][:, col_indices]
        
        ids1_new = [self.ids1[idx] for idx in row_indices]
        ids2_new = [self.ids2[idx] for idx in col_indices]
        
        if ids1_new and ids2_new:
            new_instance = GeneralMatrix(ids1_new, ids2_new)
            new_instance.matrix = submatrix
            return new_instance
        else:
            return False
        
    def Calc_distances(self, coords1, coords2):
        """
        Calculate mutual distances between two lists of 3D coordinates and fill the matrix.
        
        Parameters:
        coords1 (list of np.array): List of 3D coordinates for ids1.
        coords2 (list of np.array): List of 3D coordinates for ids2.
        """
        if len(coords1) != len(self.ids1) or len(coords2) != len(self.ids2):
            raise ValueError("The length of coordinates lists must match the length of ID lists.")
        
        coords1 = np.array(coords1)
        coords2 = np.array(coords2)
        # Vectorized calculation of distances
        distances = np.linalg.norm(coords1[:, np.newaxis, :] - coords2[np.newaxis, :, :], axis=2)
        distance_matrix = np.where(distances is None, np.nan, distances)
        
        for i in range(len(self.ids1)):
            for j in range(len(self.ids2)):
                self.matrix[i, j] = distance_matrix[i, j]
                
    def Calc_average(self):
      return np.nanmean(self.matrix)

class ExcelTable():
  def __init__(self, file_path = ""):
    self.path = file_path
    if file_path == "" or not os.path.exists(file_path):
      self.DF = pd.DataFrame()  # Initialize an empty DataFrame
      print("Initialized an empty DataFrame for EXCEL")
    else:
      self.DF = pd.read_excel(file_path, index_col=None, usecols=lambda x: x != 0)
    
    self.redund_idxs = {}
    self.unique_idxs = {}
    
  def row_iterator(self):
    return self.DF.iterrows()
  
  def Index(self, field, unique = True):
    if unique:
      for idx, inRow in self.row_iterator():
        ID = inRow[field]
        if ID not in self.unique_idxs: 
          self.unique_idxs[ID] = idx
        else:
          print(f"Something is wrong. The row with ID {ID} does not seem to be unique!")
    else:
      for idx, inRow in self.row_iterator():
        ID = inRow[field]
        if ID not in self.redund_idxs: 
          self.redund_idxs[ID] = []
        self.redund_idxs[ID].append(idx)

  def Add_path(self, file_path):
    self.path = file_path

  def Update_row(self, idx, update_row):
    for key, value in update_row.items():
      if key in self.DF.columns:
        self.DF.at[idx, key] = value
      else:
        print(f"There is no Column named '{key}'")
        
  def Append_row(self, app_row):
    self.DF.loc[len(self.DF)] = app_row

  def Append_rows(self, inp_rows):
    Ncol = len(self.DF.columns)
    if len(inp_rows) > 0:
      empty_cols = Ncol - len(inp_rows[0])
      if empty_cols > 0:
        rows = [row + [None] * empty_cols for row in inp_rows]
        #rows = [row + [''] * empty_cols for row in inp_rows]
      else:
        rows = [row[:Ncol] for row in inp_rows]
      
      excel_List = self.DF.values.tolist()
      print("BEFORE ", len(excel_List))
      for row in rows:
        #print("*"*99)
        #print(row)
        #print("-"*99)
        excel_List.append(row)
        
      print("AFTER ", len(excel_List))
      #outDF = pd.DataFrame(excel_List, columns = self.DF.columns)
      #outDF.to_excel(self.path, index=False)

      self.DF = pd.DataFrame(excel_List, columns=self.DF.columns)
      self.DF.to_excel(self.path, index=False)
      
    else:
      print("Input Rows are empty")

  def Add_empty_column(self, column_name, also_save = True):
    if column_name not in self.DF.columns:
      self.DF[column_name] = None
    else:
      print(f"Column '{column_name}' already exists in the DataFrame.")
    
    if also_save:
      if self.path:
        self.DF.to_excel(self.path, index=False)

  def Add_empty_columns(self, column_names, with_save = True):
    for column_name in column_names:
      self.Add_empty_column(column_name, with_save)

  def Add_columns(self, columns):
    if isinstance(columns, dict):
      # Add new columns with default values from the dictionary
      for col_name, default_value in columns.items():
        if col_name not in self.DF.columns:
          self.DF[col_name] = default_value
    elif isinstance(columns, list):
      # Add new columns with None values
      for col_name in columns:
        if col_name not in self.DF.columns:
          self.DF[col_name] = None
    else:
      raise TypeError("The 'columns' parameter must be either a dictionary or a list.")

    if self.path:
      self.DF.to_excel(self.path, index=False)
  
  def Set_path(self, path):
    if path != '':
      self.path = path
  
  def Just_save(self):  
    if self.path:
      try:
        self.DF.to_excel(self.path, index=False)
        #print(f"DataFrame saved successfully to {self.path}")
      except PermissionError:
        new_path = modify_filename(self.path)
        self.DF.to_excel(new_path, index=False)
        print(f"Error: The file {self.path} is open. Saved to a new file: {new_path}")
      except OSError as e:
        print(f"Error: Unable to save the file to {self.path}. Details: {e}")

  def Save(self, path):  
    if path != '':
      self.path = path
      #self.DF.to_excel(self.path, index=False)
      self.Just_save()
      
      
      
      
      
      